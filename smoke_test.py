"""전체 흐름 스모크 테스트: 관리자 생성 → 기기 등록 → 출퇴근 → 급여 → 역산 → 엑셀."""
import os, re, sys
os.environ["DATABASE_URL"] = "sqlite:///./smoke.db"
os.environ["SECRET_KEY"] = "test-key"
if os.path.exists("smoke.db"):
    os.remove("smoke.db")

from fastapi.testclient import TestClient
from sqlmodel import Session, select

from app.main import app
from app.db import engine, init_db
from app.models import AdminUser, Record, Worker
from app.security import hash_secret
from app.payroll import gross_from_net, withhold, calc_span, legal_break

init_db()
with Session(engine) as db:
    db.add(AdminUser(username="owner", password_hash=hash_secret("pw1234"), role="owner"))
    db.commit()

admin = TestClient(app)
r = admin.post("/admin/login", data={"username": "owner", "password": "pw1234"}, follow_redirects=False)
assert r.status_code == 303, r.status_code
print("✓ 관리자 로그인")

# 단말기 인증 코드 확인
r = admin.get("/admin/settings")
assert r.status_code == 200
r = admin.post("/admin/settings/passcode", data={"passcode": "481502"}, follow_redirects=False)
assert r.status_code == 303
print("✓ 단말기 인증 코드 설정")

# 직원 등록
r = admin.post("/admin/workers", data={"name": "김파트", "pin": "1234", "hourly": "10030",
                                       "contract_days": "5", "employment_type": "파트타임",
                                       "break_policy": "ask", "weekly_holiday_policy": "auto",
                                       "net_pay_agreement": "true"}, follow_redirects=False)
assert r.status_code == 303
print("✓ 직원 등록")

# ── POS PC 흉내: 별도 클라이언트 ──
pos = TestClient(app)
r = pos.get("/", headers={"accept": "text/html"}, follow_redirects=False)
assert r.status_code == 303 and r.headers["location"] == "/kiosk-login", r.status_code
print("✓ 미인증 기기 → 인증 코드 페이지로 리다이렉트")

r = pos.post("/kiosk-login", data={"passcode": "0000"}, follow_redirects=False)
assert r.status_code == 401
print("✓ 잘못된 인증 코드 거부")
r = pos.post("/kiosk-login", data={"passcode": "481502"}, follow_redirects=False)
assert r.status_code == 303, r.status_code
r = pos.get("/")
assert r.status_code == 200 and "번호를 입력해 주세요" in r.text
print("✓ 등록 후 키오스크 정상 표시")

# 관리자 브라우저에서는 키오스크가 막혀야 함 (기기 쿠키 없음)
assert admin.get("/", headers={"accept": "text/html"},
                 follow_redirects=False).status_code == 303
# 반대로 POS PC 에서 관리자 페이지는 로그인 요구
assert pos.get("/admin/payroll", headers={"accept": "text/html"},
               follow_redirects=False).status_code == 303
print("✓ 근로자/관리자 라우팅 분리 확인")

# 출퇴근
r = pos.post("/api/login", json={"pin": "9999"})
assert r.status_code == 404
r = pos.post("/api/login", json={"pin": "1234"})
assert r.status_code == 200 and r.json()["name"] == "김파트"
r = pos.post("/api/clock-in")
assert r.status_code == 200, r.text
print("✓ 출근:", r.json()["in_time"])

# 근로자 API 는 금액을 절대 내보내지 않아야 함
body = r.json()
assert not any(k in body for k in ("hourly", "total_pay", "base_pay")), body
print("✓ 근로자 응답에 급여 정보 없음")

r = pos.post("/api/clock-out", json={"break_minutes": 0})
assert r.status_code == 200, r.text
out = r.json()
print(f"✓ 퇴근: 휴게 {out['break_minutes']}분 / 실근무 {out['work_minutes']}분")
assert out["break_minutes"] == 0, "휴게 0분 선택이 반영되지 않음"

r = pos.get("/api/my-records")
assert r.status_code == 200 and r.json()["days"] == 1
print("✓ 본인 기록 조회")

# 과거 기록 주입 (주휴 요건 충족용)
from datetime import date, timedelta
with Session(engine) as db:
    w = db.exec(select(Worker).where(Worker.name == "김파트")).first()
    base = date.today() - timedelta(days=30)
    monday = base - timedelta(days=base.weekday())
    for i in range(5):
        d = monday + timedelta(days=i)
        db.add(Record(worker_id=w.id, work_date=d, in_h=9, in_m=0, out_h=18, out_m=0,
                      total_minutes=540, break_minutes=60, break_override=None,
                      break_source="auto", minutes=480))
    db.commit()
    y, m = monday.year, monday.month
    wid = w.id

r = admin.get(f"/admin/payroll?year={y}&month={m}")
assert r.status_code == 200, r.text
print("✓ 급여 대시보드 렌더링")

r = admin.post("/admin/payroll/adjust", data={"worker_id": wid, "year": y, "month": m,
                                              "amount": "50000", "memo": "교통비"})
assert r.status_code == 200 and 'value="50000"' in r.text
print("✓ 화면에서 조정액 수정 저장")

r = admin.post("/admin/payroll/grossup", data={"year": y, "month": m, "worker_ids": [wid]})
assert r.status_code == 200 and "세전" in r.text
print("✓ 3.3% 역산 패널 (세후합의 대상)")

# 세후합의가 아닌 직원은 역산에서 제외돼야 함
with Session(engine) as db:
    w2 = Worker(name="박정규", pin_hash=hash_secret("5678"), hourly=12000,
                employment_type="정규", net_pay_agreement=False)
    db.add(w2); db.commit(); db.refresh(w2)
    base2 = date.today() - timedelta(days=30)
    mon2 = base2 - timedelta(days=base2.weekday())
    for i in range(5):
        db.add(Record(worker_id=w2.id, work_date=mon2 + timedelta(days=i),
                      in_h=9, in_m=0, out_h=18, out_m=0, total_minutes=540,
                      break_minutes=60, break_source="auto", minutes=480))
    db.commit()
    wid2 = w2.id

r = admin.post("/admin/payroll/grossup",
               data={"year": y, "month": m, "worker_ids": [wid, wid2]})
assert "역산에서 제외된" in r.text and "박정규" in r.text, r.text[:600]
print("✓ 세후합의 아닌 직원은 역산 제외됨")

r = admin.post("/admin/payroll/export", data={"year": y, "month": m,
                                              "worker_ids": [wid, wid2],
                                              "include_grossup": "true"})
assert r.status_code == 200 and r.headers["content-type"].startswith(
    "application/vnd.openxml"), r.headers
open("/tmp/out.xlsx", "wb").write(r.content)
print(f"✓ 선택 인원 엑셀 내보내기 ({len(r.content):,} bytes)")

# 엑셀에서도 비대상 직원이 역산되지 않았는지 확인
import openpyxl
from io import BytesIO
ws = openpyxl.load_workbook(BytesIO(r.content)).active
rows_x = {ws.cell(row=i, column=2).value: (ws.cell(row=i, column=10).value,
                                           ws.cell(row=i, column=11).value,
                                           ws.cell(row=i, column=12).value)
          for i in (3, 4)}
net_p, gross_p, tax_p = rows_x["김파트"]
net_r, gross_r, tax_r = rows_x["박정규"]
assert gross_p > net_p and tax_p != "-", f"세후합의 직원 역산 안 됨: {rows_x['김파트']}"
assert gross_r == net_r and tax_r == "-", f"정규직에 역산 적용됨: {rows_x['박정규']}"
print(f"  · 김파트(세후합의): {net_p:,} → 세전 {gross_p:,}")
print(f"  · 박정규(정규직)  : {net_r:,} → 그대로 {gross_r:,} (역산 미적용)")

r = admin.post("/admin/payroll/export", data={"year": y, "month": m})
assert r.status_code == 400
print("✓ 미선택 시 내보내기 거부")

# 주휴 토글
r = admin.post("/admin/weekly-holiday", data={"worker_id": wid, "iso_year": y,
                                              "iso_week": monday.isocalendar()[1],
                                              "granted": "false"})
assert "미지급" in r.text
print("✓ 주휴수당 주 단위 수동 토글")

r = admin.get(f"/admin/records?year={y}&month={m}")
assert r.status_code == 200
with Session(engine) as db:
    rec = db.exec(select(Record).where(Record.worker_id == wid)).all()[0]
r = admin.post(f"/admin/records/{rec.id}", data={"in_h": 9, "in_m": 0, "out_h": 18, "out_m": 0,
                                                 "break_mode": "manual", "break_minutes": 0,
                                                 "memo": "휴게 못함"})
assert r.status_code == 200 and "관리자" in r.text
print("✓ 관리자 휴게시간 수동 override")

assert admin.get("/admin/workers").status_code == 200
print("✓ 직원 관리 페이지")

# ── 순수 로직 검증 ──
print("\n[역산 검증]")
for net in (1_000_000, 967_000, 500_000, 1_234_567, 87_000):
    g = gross_from_net(net)
    back = withhold(g["gross"])
    flag = "일치" if back["net"] == net else f"근사({back['net']:,})"
    print(f"  세후 {net:>9,}원 → 세전 {g['gross']:>9,}원 "
          f"(소득세 {g['income_tax']:,} + 지방 {g['local_tax']:,}) {flag}")
    assert abs(back["net"] - net) <= 10

print("\n[휴게시간 override 검증]")
s_auto = calc_span(9, 0, 18, 0)
s_none = calc_span(9, 0, 18, 0, break_override=0)
s_manu = calc_span(9, 0, 18, 0, break_override=20)
print(f"  자동   : 체류 {s_auto.total_min}분, 휴게 {s_auto.break_min}분 → 실근무 {s_auto.paid_min}분")
print(f"  안 쉼  : 휴게 {s_none.break_min}분 → 실근무 {s_none.paid_min}분 "
      f"(법정 미만 경고={s_none.break_is_short})")
print(f"  20분   : 휴게 {s_manu.break_min}분 → 실근무 {s_manu.paid_min}분")
assert (s_auto.paid_min, s_none.paid_min, s_manu.paid_min) == (480, 540, 520)

night = calc_span(22, 0, 6, 0, break_override=30)
print(f"  야간조 : 22:00→06:00 = {night.total_min}분, 실근무 {night.paid_min}분")
assert night.total_min == 480

print("\n전부 통과 ✅")
