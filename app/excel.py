"""선택된 인원만 담는 엑셀 생성 (요구사항 5-②).

report.py 의 서식을 그대로 옮기되, 전체 직원이 아니라 화면에서 체크한
인원 목록(rows)만 받도록 바꿨습니다. 파일로 저장하지 않고 메모리에서
바로 스트리밍하므로 서버에 파일이 쌓이지 않습니다.
"""
from __future__ import annotations

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .payroll import gross_from_net

HEADER_FILL = PatternFill("solid", fgColor="69B5C2")
SUB_FILL = PatternFill("solid", fgColor="B0D4DA")
MONEY = '#,##0"원"'
CENTER = Alignment(horizontal="center", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def build_payroll_workbook(rows: list[dict], year: int, month: int,
                           include_grossup: bool = False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}년{month:02d}월"

    headers = ["No.", "이름", "구분", "시급", "실근무",
               "기본급", "주휴수당", "가산수당", "조정액", "지급액"]
    if include_grossup:
        headers += ["세전(3.3% 역산)", "소득세", "지방소득세"]

    last_col = len(headers)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_col)
    title = ws.cell(row=1, column=1, value=f"{year}년 {month}월 급여지급 리스트")
    title.font = Font(name="맑은 고딕", bold=True, size=13, color="FFFFFF")
    title.fill = HEADER_FILL
    title.alignment = CENTER
    ws.row_dimensions[1].height = 28

    for col, text in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=text)
        c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        c.fill = SUB_FILL
        c.alignment = CENTER
    ws.row_dimensions[2].height = 20

    money_cols = {4, 6, 7, 8, 9, 10, 11, 12, 13}
    total_net = total_gross = 0

    for i, r in enumerate(rows):
        row_no = i + 3
        bg = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else "F0F4F8")
        vals = [i + 1, r["name"], r["employment_type"], r["hourly"],
                f"{r['total_hours']}h", r["base_pay"], r["weekly_pay"],
                r["extra_pay"], r["adjustment"], r["total_pay"]]

        if include_grossup:
            # 역산은 '세후합의' 로 지정된 직원에게만 적용합니다.
            # 정규직처럼 대상이 아닌 사람은 원래 지급액을 그대로 두고,
            # 세액 칸은 비워서 합계가 부풀지 않게 합니다.
            if r.get("net_pay_agreement"):
                g = gross_from_net(r["total_pay"])
                vals += [g["gross"], g["income_tax"], g["local_tax"]]
                total_gross += g["gross"]
            else:
                vals += [r["total_pay"], "-", "-"]
                total_gross += r["total_pay"]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=row_no, column=col, value=val)
            c.fill = bg
            c.font = Font(name="맑은 고딕", size=10)
            c.alignment = CENTER
            if col in money_cols:
                c.number_format = MONEY
        total_net += r["total_pay"]

    # 합계 행
    tr = len(rows) + 3
    ws.merge_cells(start_row=tr, start_column=1, end_row=tr, end_column=9)
    c = ws.cell(row=tr, column=1, value=f"선택 {len(rows)}명 합계")
    c.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    c.fill = HEADER_FILL
    c.alignment = RIGHT

    c2 = ws.cell(row=tr, column=10, value=total_net)
    c2.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
    c2.fill = HEADER_FILL
    c2.number_format = MONEY
    c2.alignment = CENTER

    if include_grossup:
        c3 = ws.cell(row=tr, column=11, value=total_gross)
        c3.font = Font(name="맑은 고딕", bold=True, size=10, color="FFFFFF")
        c3.fill = HEADER_FILL
        c3.number_format = MONEY
        c3.alignment = CENTER
        ws.cell(row=tr + 2, column=1,
                value="※ 3.3% 역산액은 참고용입니다. 실제 신고 전 세무대리인 확인이 필요합니다.")

    widths = [5, 12, 9, 10, 9, 12, 12, 12, 10, 14, 14, 12, 12]
    for col in range(1, last_col + 1):
        ws.column_dimensions[get_column_letter(col)].width = widths[col - 1]

    ws.freeze_panes = "A3"
    return wb
